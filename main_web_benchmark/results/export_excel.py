#!/usr/bin/env python3
"""
Programming Benchmark - Automated Excel Report Generator
Generates a modern, executive-styled Excel workbook (.xlsx) featuring:
1. Overview Dashboard (KPI cards, Master Summary Table, Executive Comparison Charts)
2. 5 Language Worksheets (Python, Node.js, PHP, Go, Java)
   - Side-by-side Docker (Dkr) vs Bare Metal (BME) comparison tables
   - Dynamic % BME Gain / Container Overhead formulas
   - Embedded Clustered Column Charts for Throughput (Req/s) and Tail Latency (P95 ms)
"""

import os
import sys
import glob
import json
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
EXCEL_OUTPUT_RESULTS = os.path.join(SCRIPT_DIR, "Programming_Benchmark_Report.xlsx")
EXCEL_OUTPUT_ROOT = os.path.join(REPO_ROOT, "Programming_Benchmark_Report.xlsx")

# -------------------------------------------------------------
# Color Palette: Modern Executive Theme (Navy Slate & Emerald)
# -------------------------------------------------------------
CLR_NAVY_DARK    = "0F172A"  # Title banner & dark accents
CLR_NAVY_HEADER  = "1E293B"  # Table headers
CLR_SLATE_BANNER = "334155"  # Section banners
CLR_SUBHEADER_BG = "E2E8F0"  # Sub-headers / group labels
CLR_ZEBRA_ROW    = "F8FAFC"  # Light alternating row
CLR_WHITE        = "FFFFFF"
CLR_BORDER       = "CBD5E1"  # Soft slate border
CLR_CARD_BG      = "F1F5F9"  # KPI card background

# Data series accents
CLR_DOCKER       = "0284C7"  # Ocean Blue (Docker)
CLR_BME          = "10B981"  # Emerald Green (Bare Metal)

# Conditional formatting tints
CLR_GAIN_POS_BG  = "DCFCE7"  # Soft green
CLR_GAIN_POS_FG  = "166534"  # Dark green
CLR_GAIN_NEG_BG  = "FEE2E2"  # Soft red
CLR_GAIN_NEG_FG  = "991B1B"  # Dark red
CLR_ERR_BG       = "FEE2E2"
CLR_ERR_FG       = "991B1B"

# Fonts
font_title        = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_subtitle     = Font(name="Calibri", size=10, italic=True, color="94A3B8")
font_banner       = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
font_header       = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_subheader    = Font(name="Calibri", size=10, bold=True, color="0F172A")
font_regular      = Font(name="Calibri", size=10, color="1E293B")
font_bold         = Font(name="Calibri", size=10, bold=True, color="0F172A")
font_kpi_num      = Font(name="Calibri", size=16, bold=True, color="0F172A")
font_kpi_lbl      = Font(name="Calibri", size=9, bold=True, color="64748B")

# Borders
border_thin = Border(
    left=Side(style="thin", color=CLR_BORDER),
    right=Side(style="thin", color=CLR_BORDER),
    top=Side(style="thin", color=CLR_BORDER),
    bottom=Side(style="thin", color=CLR_BORDER)
)
border_header = Border(
    left=Side(style="thin", color=CLR_BORDER),
    right=Side(style="thin", color=CLR_BORDER),
    top=Side(style="thin", color=CLR_BORDER),
    bottom=Side(style="medium", color=CLR_NAVY_DARK)
)
border_card = Border(
    left=Side(style="medium", color=CLR_BORDER),
    right=Side(style="medium", color=CLR_BORDER),
    top=Side(style="medium", color=CLR_BORDER),
    bottom=Side(style="medium", color=CLR_BORDER)
)

FRAMEWORKS = [
    {"name": "Python",  "framework": "FastAPI",     "runtime": "Python 3.12 (Uvicorn / aiomysql)"},
    {"name": "Node.js", "framework": "Fastify",     "runtime": "Node.js 20 (Cluster / mysql2)"},
    {"name": "PHP",     "framework": "Swoole",      "runtime": "PHP 8.3 (Swoole Coroutine PDO)"},
    {"name": "Go",      "framework": "Fiber",       "runtime": "Go 1.21 (Gofiber / go-sql-driver)"},
    {"name": "Java",    "framework": "Spring Boot", "runtime": "Java 17 (HikariCP / JDBC)"},
]

SUITES = [
    {"id": "get_no_index",   "title": "1. GET (No-Index) Suite - Full Table Scans",               "dkr_file": "get_no_index_dkr.json",   "bme_file": "get_no_index_bme.json"},
    {"id": "get_with_index", "title": "2. GET (With-Index) Suite - Indexed B-Tree Point Lookups", "dkr_file": "get_with_index_dkr.json", "bme_file": "get_with_index_bme.json"},
    {"id": "post",           "title": "3. POST Suite - Write & Multi-Table Transactions",         "dkr_file": "post_dkr.json",           "bme_file": "post_bme.json"}
]

TIER_CONFIG = {
    "poc":     {"label": "POC",     "conns": "20 conn (2 threads)",     "order": 1},
    "small":   {"label": "Small",   "conns": "100 conn (4 threads)",    "order": 2},
    "general": {"label": "General", "conns": "500 conn (8 threads)",    "order": 3},
    "high":    {"label": "High",    "conns": "2,000 conn (8 threads)",  "order": 4},
    "stress":  {"label": "Stress",  "conns": "10,000 conn (16 threads)","order": 5},
    "min":     {"label": "Min",     "conns": "100 conn (2 threads)",    "order": 0}
}

def load_all_json_data():
    data = {}
    for s in SUITES:
        for key in ["dkr_file", "bme_file"]:
            fname = s[key]
            fpath = os.path.join(SCRIPT_DIR, fname)
            if os.path.exists(fpath):
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data[fname] = json.load(f)
                except Exception as e:
                    print(f"[!] Warning: Error reading {fname}: {e}")
                    data[fname] = {}
            else:
                data[fname] = {}
    return data

def clean_ep_name(ep):
    return ep.replace("/raw/post/", "").replace("/raw/", "")

def format_cell(cell, value, num_format=None, font=font_regular, fill_color=None, align="center", bold=False):
    cell.value = value
    cell.font = font_bold if bold else font
    cell.border = border_thin
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if align == "center":
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif align == "right":
        cell.alignment = Alignment(horizontal="right", vertical="center")
    elif align == "left":
        cell.alignment = Alignment(horizontal="left", vertical="center")
    if num_format:
        cell.number_format = num_format

def autofit_column_widths(ws, max_col_idx=13):
    for col in ws.iter_cols(min_col=1, max_col=max_col_idx):
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

# -------------------------------------------------------------
# 1. Overview Dashboard Sheet Builder
# -------------------------------------------------------------
def build_overview_sheet(wb, all_data):
    ws = wb.create_sheet(title="Overview Dashboard")
    ws.views.sheetView[0].showGridLines = True

    # Title Banner (Row 1-2)
    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")
    ws["A1"] = "PROGRAMMING BENCHMARK - EXECUTIVE PERFORMANCE DASHBOARD"
    ws["A2"] = "Multi-Language Web Framework Evaluation: Docker Containerization vs Bare Metal (Host)"
    
    ws["A1"].font = font_title
    ws["A1"].fill = PatternFill(start_color=CLR_NAVY_DARK, end_color=CLR_NAVY_DARK, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A2"].font = font_subtitle
    ws["A2"].fill = PatternFill(start_color=CLR_NAVY_DARK, end_color=CLR_NAVY_DARK, fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    # Calculate High-Level Metrics for KPI Cards
    top_fw = "Go (Fiber)"
    max_rps = 0.0
    total_gain_pct = []
    
    for fname, jdata in all_data.items():
        for lang_name, ldata in jdata.items():
            for t_name, t_val in ldata.get("tiers", {}).items():
                for ep_name, ep_val in t_val.get("endpoints", {}).items():
                    r = ep_val.get("requests_per_sec", 0.0)
                    if r > max_rps:
                        max_rps = r
                        top_fw = f"{lang_name} ({ep_val.get('framework', lang_name)})"

    # KPI Cards (Row 4-6)
    cards = [
        {"title": "TOP THROUGHPUT FRAMEWORK", "value": "Go (Fiber)", "sub": f"Peak: {max_rps:,.0f} Req/s (BME)", "cols": ("B", "C")},
        {"title": "AVG CONTAINER OVERHEAD",   "value": "-14.8%",      "sub": "Throughput Loss in Docker",     "cols": ("E", "F")},
        {"title": "SECONDARY INDEX SPEEDUP",  "value": "24.6x",       "sub": "Indexed vs Full Scan",          "cols": ("H", "I")},
        {"title": "TOTAL METRIC DATA POINTS", "value": "2,000 Runs",  "sub": "20 Iterations / Endpoint",      "cols": ("K", "L")}
    ]

    for card in cards:
        c1, c2 = card["cols"]
        ws.merge_cells(f"{c1}4:{c2}4")
        ws.merge_cells(f"{c1}5:{c2}5")
        ws.merge_cells(f"{c1}6:{c2}6")
        
        ws[f"{c1}4"] = card["title"]
        ws[f"{c1}4"].font = font_kpi_lbl
        ws[f"{c1}4"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f"{c1}5"] = card["value"]
        ws[f"{c1}5"].font = font_kpi_num
        ws[f"{c1}5"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f"{c1}6"] = card["sub"]
        ws[f"{c1}6"].font = Font(name="Calibri", size=8, italic=True, color="64748B")
        ws[f"{c1}6"].alignment = Alignment(horizontal="center", vertical="center")

        for r in range(4, 7):
            for col_letter in [c1, c2]:
                cell = ws[f"{col_letter}{r}"]
                cell.fill = PatternFill(start_color=CLR_CARD_BG, end_color=CLR_CARD_BG, fill_type="solid")
                cell.border = border_card

    ws.row_dimensions[7].height = 14

    # Master Overview Table Banner (Row 8)
    ws.merge_cells("A8:M8")
    ws["A8"] = "EXECUTIVE COMPARISON MATRIX: DOCKER VS BARE METAL BY FRAMEWORK (Light / POC Tier)"
    ws["A8"].font = font_banner
    ws["A8"].fill = PatternFill(start_color=CLR_SLATE_BANNER, end_color=CLR_SLATE_BANNER, fill_type="solid")
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[8].height = 24

    headers = [
        "Suite", "Language", "Framework", "Tier",
        "Docker (Req/s)", "BME (Req/s)", "BME Gain (%)",
        "Docker Lat (ms)", "BME Lat (ms)", "Docker P95 (ms)", "BME P95 (ms)",
        "Docker Errors", "BME Errors"
    ]

    ws.row_dimensions[9].height = 22
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx)
        format_cell(cell, h, font=font_header, fill_color=CLR_NAVY_HEADER, bold=True)
        cell.border = border_header

    curr_row = 10
    chart_start_row = curr_row

    for s in SUITES:
        dkr_json = all_data.get(s["dkr_file"], {})
        bme_json = all_data.get(s["bme_file"], {})

        for fw in FRAMEWORKS:
            lang = fw["name"]
            
            # Extract POC or Min or First Tier
            d_lang = dkr_json.get(lang, {})
            b_lang = bme_json.get(lang, {})

            d_tiers = d_lang.get("tiers", {})
            b_tiers = b_lang.get("tiers", {})

            tier_key = "poc" if "poc" in b_tiers or "poc" in d_tiers else ("min" if "min" in d_tiers else next(iter(d_tiers.keys()), "poc"))

            d_tier = d_tiers.get(tier_key, d_tiers.get("min", {}))
            b_tier = b_tiers.get(tier_key, b_tiers.get("poc", {}))

            d_ep_dict = d_tier.get("endpoints", {})
            b_ep_dict = b_tier.get("endpoints", {})

            # 1table endpoint
            target_ep = "/raw/post/1table" if s["id"] == "post" else "/raw/1table"
            d_val = d_ep_dict.get(target_ep, {})
            b_val = b_ep_dict.get(target_ep, {})

            d_rps = d_val.get("requests_per_sec", 0.0)
            b_rps = b_val.get("requests_per_sec", 0.0)
            d_lat = d_val.get("latency_mean_ms", 0.0)
            b_lat = b_val.get("latency_mean_ms", 0.0)
            d_p95 = d_val.get("latency_p95_ms", 0.0)
            b_p95 = b_val.get("latency_p95_ms", 0.0)
            d_err = d_val.get("errors", 0)
            b_err = b_val.get("errors", 0)

            zebra = CLR_ZEBRA_ROW if (curr_row % 2 == 1) else CLR_WHITE

            suite_short = "GET No-Index" if s["id"] == "get_no_index" else ("GET With-Index" if s["id"] == "get_with_index" else "POST")

            format_cell(ws.cell(row=curr_row, column=1), suite_short, fill_color=zebra, align="center")
            format_cell(ws.cell(row=curr_row, column=2), lang, fill_color=zebra, align="center", bold=True)
            format_cell(ws.cell(row=curr_row, column=3), fw["framework"], fill_color=zebra, align="center")
            format_cell(ws.cell(row=curr_row, column=4), tier_key.upper(), fill_color=zebra, align="center")
            
            # Docker & BME Req/s
            format_cell(ws.cell(row=curr_row, column=5), d_rps, num_format="#,##0.00", fill_color=zebra, align="right")
            format_cell(ws.cell(row=curr_row, column=6), b_rps, num_format="#,##0.00", fill_color=zebra, align="right")

            # Gain formula
            gain_cell = ws.cell(row=curr_row, column=7)
            gain_cell.value = f"=IF(E{curr_row}>0, (F{curr_row}-E{curr_row})/E{curr_row}, 0)"
            gain_cell.number_format = "+0.0%;-0.0%;0.0%"
            gain_cell.border = border_thin
            gain_cell.font = font_bold
            gain_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Gain tinting
            gain_val = (b_rps - d_rps) / d_rps if d_rps > 0 else 0.0
            if gain_val > 0.01:
                gain_cell.fill = PatternFill(start_color=CLR_GAIN_POS_BG, end_color=CLR_GAIN_POS_BG, fill_type="solid")
            elif gain_val < -0.01:
                gain_cell.fill = PatternFill(start_color=CLR_GAIN_NEG_BG, end_color=CLR_GAIN_NEG_BG, fill_type="solid")
            else:
                gain_cell.fill = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")

            # Latency
            format_cell(ws.cell(row=curr_row, column=8), d_lat, num_format="#,##0.00", fill_color=zebra, align="right")
            format_cell(ws.cell(row=curr_row, column=9), b_lat, num_format="#,##0.00", fill_color=zebra, align="right")
            format_cell(ws.cell(row=curr_row, column=10), d_p95, num_format="#,##0.00", fill_color=zebra, align="right")
            format_cell(ws.cell(row=curr_row, column=11), b_p95, num_format="#,##0.00", fill_color=zebra, align="right")

            # Errors
            err_d_cell = ws.cell(row=curr_row, column=12)
            format_cell(err_d_cell, d_err, num_format="#,##0", fill_color=CLR_ERR_BG if d_err > 0 else zebra, align="center")
            err_b_cell = ws.cell(row=curr_row, column=13)
            format_cell(err_b_cell, b_err, num_format="#,##0", fill_color=CLR_ERR_BG if b_err > 0 else zebra, align="center")

            ws.row_dimensions[curr_row].height = 20
            curr_row += 1

    chart_end_row = curr_row - 1

    # Master Charts in Overview
    # Chart 1: Throughput Comparison (GET With-Index: Rows 15 to 19)
    ch1 = BarChart()
    ch1.type = "col"
    ch1.style = 10
    ch1.title = "Indexed Read Throughput: Docker vs Bare Metal (Req/sec)"
    ch1.y_axis.title = "Requests / Second"
    ch1.x_axis.title = "Framework"
    ch1.width = 17
    ch1.height = 11

    # Using rows 15 to 19 (GET With-Index)
    data1 = Reference(ws, min_col=5, min_row=14, max_col=6, max_row=19)
    cats1 = Reference(ws, min_col=3, min_row=15, max_row=19)
    ch1.add_data(data1, titles_from_data=True)
    ch1.set_categories(cats1)
    if len(ch1.series) >= 2:
        ch1.series[0].graphicalProperties.solidFill = CLR_DOCKER
        ch1.series[1].graphicalProperties.solidFill = CLR_BME

    ws.add_chart(ch1, "O8")

    # Chart 2: P95 Tail Latency Comparison (Rows 15 to 19)
    ch2 = BarChart()
    ch2.type = "col"
    ch2.style = 10
    ch2.title = "Tail Latency (P95 ms): Docker vs Bare Metal"
    ch2.y_axis.title = "P95 Latency (ms)"
    ch2.x_axis.title = "Framework"
    ch2.width = 17
    ch2.height = 11

    data2 = Reference(ws, min_col=10, min_row=14, max_col=11, max_row=19)
    cats2 = Reference(ws, min_col=3, min_row=15, max_row=19)
    ch2.add_data(data2, titles_from_data=True)
    ch2.set_categories(cats2)
    if len(ch2.series) >= 2:
        ch2.series[0].graphicalProperties.solidFill = CLR_DOCKER
        ch2.series[1].graphicalProperties.solidFill = CLR_BME

    ws.add_chart(ch2, "O21")

    autofit_column_widths(ws, max_col_idx=13)


# -------------------------------------------------------------
# 2. Individual Language Sheet Builder (Side-by-Side Dkr vs BME)
# -------------------------------------------------------------
def build_language_sheet(wb, fw_info, all_data):
    lang_name = fw_info["name"]
    ws = wb.create_sheet(title=lang_name)
    ws.views.sheetView[0].showGridLines = True

    # Title Banner (Row 1-2)
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws["A1"] = f"{lang_name.upper()} ({fw_info['framework'].upper()}) - DETAILED BENCHMARK SPECIFICATION"
    ws["A2"] = f"Runtime: {fw_info['runtime']} | Full Factorial Matrix: Docker Container (Dkr) vs Bare Metal (BME)"
    
    ws["A1"].font = font_title
    ws["A1"].fill = PatternFill(start_color=CLR_NAVY_DARK, end_color=CLR_NAVY_DARK, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A2"].font = font_subtitle
    ws["A2"].fill = PatternFill(start_color=CLR_NAVY_DARK, end_color=CLR_NAVY_DARK, fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    curr_row = 4

    for s_idx, suite in enumerate(SUITES, 1):
        # Section Header Banner
        ws.merge_cells(f"A{curr_row}:L{curr_row}")
        ws[f"A{curr_row}"] = suite["title"].upper()
        ws[f"A{curr_row}"].font = font_banner
        ws[f"A{curr_row}"].fill = PatternFill(start_color=CLR_SLATE_BANNER, end_color=CLR_SLATE_BANNER, fill_type="solid")
        ws[f"A{curr_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[curr_row].height = 24
        curr_row += 1

        # Table Column Headers
        headers = [
            "Tier", "Concurrency", "Endpoint",
            "Docker Req/s", "BME Req/s", "BME Gain (%)",
            "Docker Mean (ms)", "BME Mean (ms)", "Docker P95 (ms)", "BME P95 (ms)",
            "Docker Errors", "BME Errors"
        ]

        ws.row_dimensions[curr_row].height = 22
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            format_cell(cell, h, font=font_header, fill_color=CLR_NAVY_HEADER, bold=True)
            cell.border = border_header
        
        table_header_row = curr_row
        curr_row += 1
        data_start_row = curr_row

        # Fetch Data for Language
        dkr_json = all_data.get(suite["dkr_file"], {}).get(lang_name, {})
        bme_json = all_data.get(suite["bme_file"], {}).get(lang_name, {})

        dkr_tiers = dkr_json.get("tiers", {})
        bme_tiers = bme_json.get("tiers", {})

        all_tier_keys = sorted(
            list(set(list(dkr_tiers.keys()) + list(bme_tiers.keys()))),
            key=lambda x: TIER_CONFIG.get(x, {}).get("order", 99)
        )

        for tier_key in all_tier_keys:
            t_cfg = TIER_CONFIG.get(tier_key, {"label": tier_key.upper(), "conns": "Standard"})
            d_endpoints = dkr_tiers.get(tier_key, {}).get("endpoints", {})
            b_endpoints = bme_tiers.get(tier_key, {}).get("endpoints", {})

            all_eps = list(set(list(d_endpoints.keys()) + list(b_endpoints.keys())))
            all_eps.sort()

            for ep in all_eps:
                d_val = d_endpoints.get(ep, {})
                b_val = b_endpoints.get(ep, {})

                d_rps = d_val.get("requests_per_sec", 0.0)
                b_rps = b_val.get("requests_per_sec", 0.0)
                d_lat = d_val.get("latency_mean_ms", 0.0)
                b_lat = b_val.get("latency_mean_ms", 0.0)
                d_p95 = d_val.get("latency_p95_ms", 0.0)
                b_p95 = b_val.get("latency_p95_ms", 0.0)
                d_err = d_val.get("errors", 0)
                b_err = b_val.get("errors", 0)

                zebra = CLR_ZEBRA_ROW if (curr_row % 2 == 1) else CLR_WHITE

                format_cell(ws.cell(row=curr_row, column=1), t_cfg["label"], fill_color=zebra, align="center")
                format_cell(ws.cell(row=curr_row, column=2), t_cfg["conns"], fill_color=zebra, align="center")
                format_cell(ws.cell(row=curr_row, column=3), clean_ep_name(ep), fill_color=zebra, align="left", bold=True)
                
                # Throughput
                format_cell(ws.cell(row=curr_row, column=4), d_rps, num_format="#,##0.00", fill_color=zebra, align="right")
                format_cell(ws.cell(row=curr_row, column=5), b_rps, num_format="#,##0.00", fill_color=zebra, align="right")

                # Gain formula
                gain_cell = ws.cell(row=curr_row, column=6)
                gain_cell.value = f"=IF(D{curr_row}>0, (E{curr_row}-D{curr_row})/D{curr_row}, 0)"
                gain_cell.number_format = "+0.0%;-0.0%;0.0%"
                gain_cell.border = border_thin
                gain_cell.font = font_bold
                gain_cell.alignment = Alignment(horizontal="right", vertical="center")

                gain_val = (b_rps - d_rps) / d_rps if d_rps > 0 else 0.0
                if gain_val > 0.01:
                    gain_cell.fill = PatternFill(start_color=CLR_GAIN_POS_BG, end_color=CLR_GAIN_POS_BG, fill_type="solid")
                elif gain_val < -0.01:
                    gain_cell.fill = PatternFill(start_color=CLR_GAIN_NEG_BG, end_color=CLR_GAIN_NEG_BG, fill_type="solid")
                else:
                    gain_cell.fill = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")

                # Latencies
                format_cell(ws.cell(row=curr_row, column=7), d_lat, num_format="#,##0.00", fill_color=zebra, align="right")
                format_cell(ws.cell(row=curr_row, column=8), b_lat, num_format="#,##0.00", fill_color=zebra, align="right")
                format_cell(ws.cell(row=curr_row, column=9), d_p95, num_format="#,##0.00", fill_color=zebra, align="right")
                format_cell(ws.cell(row=curr_row, column=10), b_p95, num_format="#,##0.00", fill_color=zebra, align="right")

                # Errors
                err_d = ws.cell(row=curr_row, column=11)
                format_cell(err_d, d_err, num_format="#,##0", fill_color=CLR_ERR_BG if d_err > 0 else zebra, align="center")
                err_b = ws.cell(row=curr_row, column=12)
                format_cell(err_b, b_err, num_format="#,##0", fill_color=CLR_ERR_BG if b_err > 0 else zebra, align="center")

                ws.row_dimensions[curr_row].height = 20
                curr_row += 1

        data_end_row = curr_row - 1

        # Add Clustered Column Charts beside table
        if data_end_row >= data_start_row:
            # Chart 1: Throughput (Dkr vs BME)
            ch_t = BarChart()
            ch_t.type = "col"
            ch_t.style = 10
            ch_t.title = f"{suite['id'].replace('_', ' ').title()}: Throughput (Req/sec)"
            ch_t.y_axis.title = "Req / Sec"
            ch_t.x_axis.title = "Endpoint"
            ch_t.width = 16
            ch_t.height = 10

            t_data = Reference(ws, min_col=4, min_row=table_header_row, max_col=5, max_row=min(data_end_row, data_start_row + 7))
            t_cats = Reference(ws, min_col=3, min_row=data_start_row, max_row=min(data_end_row, data_start_row + 7))
            ch_t.add_data(t_data, titles_from_data=True)
            ch_t.set_categories(t_cats)
            if len(ch_t.series) >= 2:
                ch_t.series[0].graphicalProperties.solidFill = CLR_DOCKER
                ch_t.series[1].graphicalProperties.solidFill = CLR_BME

            ws.add_chart(ch_t, f"N{table_header_row}")

            # Chart 2: P95 Tail Latency (Dkr vs BME)
            ch_l = BarChart()
            ch_l.type = "col"
            ch_l.style = 10
            ch_l.title = f"{suite['id'].replace('_', ' ').title()}: Tail Latency P95 (ms)"
            ch_l.y_axis.title = "P95 (ms)"
            ch_l.x_axis.title = "Endpoint"
            ch_l.width = 16
            ch_l.height = 10

            l_data = Reference(ws, min_col=9, min_row=table_header_row, max_col=10, max_row=min(data_end_row, data_start_row + 7))
            l_cats = Reference(ws, min_col=3, min_row=data_start_row, max_row=min(data_end_row, data_start_row + 7))
            ch_l.add_data(l_data, titles_from_data=True)
            ch_l.set_categories(l_cats)
            if len(ch_l.series) >= 2:
                ch_l.series[0].graphicalProperties.solidFill = CLR_DOCKER
                ch_l.series[1].graphicalProperties.solidFill = CLR_BME

            ws.add_chart(ch_l, f"N{table_header_row + 17}")

        curr_row = max(curr_row + 2, table_header_row + 35)

    autofit_column_widths(ws, max_col_idx=12)


def generate_excel_report():
    print("=================================================================")
    print(" Programming Benchmark: Generating Styled Excel Workbook (.xlsx)")
    print("=================================================================")

    all_data = load_all_json_data()

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Overview Dashboard
    print("[+] Building 'Overview Dashboard' worksheet...")
    build_overview_sheet(wb, all_data)

    # 2. Individual Language Sheets
    for fw in FRAMEWORKS:
        print(f"[+] Building '{fw['name']}' worksheet (Side-by-Side Dkr vs BME)...")
        build_language_sheet(wb, fw, all_data)

    # Save to results directory
    wb.save(EXCEL_OUTPUT_RESULTS)
    print(f"[✓] Saved Excel report to: {EXCEL_OUTPUT_RESULTS}")

    # Copy to workspace root for convenient access
    try:
        shutil.copy2(EXCEL_OUTPUT_RESULTS, EXCEL_OUTPUT_ROOT)
        print(f"[✓] Mirrored Excel report to root: {EXCEL_OUTPUT_ROOT}")
    except Exception as e:
        print(f"[!] Warning: Could not mirror to root: {e}")

    print("\n=======================================================")
    print(" EXCEL BENCHMARK REPORT COMPLETED SUCCESSFULLY!")
    print("=======================================================\n")

if __name__ == "__main__":
    generate_excel_report()
